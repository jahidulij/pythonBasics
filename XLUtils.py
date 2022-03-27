import openpyxl

def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def read_data(file, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row = rowNo, column = columnNo).value

def write_data(file, sheetName, rowNo, columnNo, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row = rowNo, column = columnNo).value = data
    workbook.save(file)

